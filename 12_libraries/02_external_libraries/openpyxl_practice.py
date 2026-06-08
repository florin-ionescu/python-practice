"""
openpyxl

What it does:
- Reads and writes Excel files.
- Works with .xlsx files.
- Useful for ticket reports, QA reports, exported data, and business reports.
"""
from openpyxl import Workbook, load_workbook

# Create a new Excel workbook
# Workbook creates a new Excel file.
workbook = Workbook()
#A sheet is one tab inside the Excel file.
sheet = workbook.active

sheet.title = "Tickets"

# Add headers
sheet["A1"] = "Ticket ID"
sheet["B1"] = "Status"
sheet["C1"] = "Priority"

# Add data
sheet.append(["INC001", "Open", "High"])
sheet.append(["INC002", "Closed", "Medium"])
sheet.append(["INC003", "In Progress", "Low"])

# Save workbook
workbook.save("tickets.xlsx")
print("Excel file created")

# Read workbook
# load_workbook opens an existing Excel file.
loaded_workbook = load_workbook('tickets.xlsx')
loaded_sheet = loaded_workbook["Tickets"]

# Read specific cell
print(loaded_sheet["A1"].value)

# Read all rows
# values_only=True means return only the cell values, not full Excel cell objects.
for row in loaded_sheet.iter_rows(values_only=True):
    print(row)